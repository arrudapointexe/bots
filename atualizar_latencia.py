import os
import glob
import zipfile
import pandas as pd
from playwright.sync_api import sync_playwright
import time
from openpyxl import load_workbook

# ==========================================
# CONFIGURAÇÕES (Adaptadas do seu shopee.py)
# ==========================================
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads_shopee")
USER_DATA_DIR = r"C:\PerfilBotShopee" # Usando o mesmo perfil para manter o login
LATENCIA_JML_PATH = os.path.join(os.getcwd(), "latenciaJML.xlsx")
LATENCIA_ITR_PATH = os.path.join(os.getcwd(), "latenciaITR.xlsx")

# ==========================================
# FUNÇÕES AUXILIARES (Reaproveitadas do seu shopee.py)
# ==========================================
def limpar_pasta_downloads():
    print("INFO: Limpando a pasta de downloads antes de começar...")
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    for arquivo in glob.glob(os.path.join(DOWNLOAD_DIR, "*")):
        try:
            os.remove(arquivo)
        except Exception as e:
            print(f"WARN: Não foi possível remover o arquivo {arquivo}: {e}")

def formatar_waybill_como_texto(caminho_arquivo):
    """
    Formata todas as colunas de waybill/tracking como texto no Excel para evitar notação científica.
    """
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        # Itera pelas colunas do header para identificar waybill/tracking
        for col_idx, col_cell in enumerate(ws[1], 1):
            if col_cell.value and any(x in str(col_cell.value).lower() for x in ['waybill', 'tracking', 'order', 'pack']):
                # Formata todas as células dessa coluna como texto
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = '@'  # @ = texto no Excel
                        if cell.value is not None:
                            cell.value = str(cell.value)  # Converte para string
        
        wb.save(caminho_arquivo)
        print(f"INFO: Arquivo '{os.path.basename(caminho_arquivo)}' formatado com waybill como texto.")
    except Exception as e:
        print(f"WARN: Erro ao formatar arquivo: {e}")

def ler_planilha_shopee_segura(caminho):
    """
    Tenta ler um arquivo CSV ou Excel com diferentes configurações para evitar erros.
    """
    print(f"INFO: Lendo o arquivo de inventário: {os.path.basename(caminho)}")
    if caminho.endswith('.csv'):
        tentativas = [
            (',', 'utf-8-sig'), (',', 'utf-8'), (',', 'latin1'),
            (';', 'utf-8-sig'), (';', 'latin1')
        ]
        for sep, enc in tentativas:
            try:
                df = pd.read_csv(caminho, sep=sep, encoding=enc, dtype=str)
                if len(df.columns) > 5:
                    print(f"SUCCESS: Arquivo CSV lido com separador '{sep}' e encoding '{enc}'.")
                    return df
            except Exception:
                continue
    elif caminho.endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(caminho, dtype=str)
            print("SUCCESS: Arquivo Excel lido com sucesso.")
            return df
        except Exception as e:
            print(f"ERROR: Falha ao ler o arquivo Excel: {e}")
    
    print("ERROR: Formato de arquivo desconhecido ou corrompido.")
    return pd.DataFrame()

# ==========================================
# 1. FUNÇÃO DE NAVEGAÇÃO E DOWNLOAD
# ==========================================
def baixar_inventario_shopee():
    """
    Navega no portal da Shopee SPX e baixa o relatório de inventário (Forward Order).
    """
    print("INFO: Iniciando o processo de download do inventário da Shopee.")
    limpar_pasta_downloads()
    
    try:
        with sync_playwright() as p:
            print("INFO: Abrindo o navegador...")
            browser = p.chromium.launch_persistent_context(
                user_data_dir=USER_DATA_DIR,
                headless=False,
                accept_downloads=True,
                slow_mo=50 # Adiciona uma pequena pausa para estabilidade
            )
            page = browser.pages[0]
            
            print("INFO: Acessando o portal SPX...")
            page.goto("https://spx.shopee.com.br/#/index", wait_until="commit", timeout=90000)
            print("INFO: Aguardando o menu principal carregar...")
            page.locator("span.sub-menu-title", has_text="Pedidos").wait_for(state="visible", timeout=90000)

            print("INFO: Navegando para 'Rastreio de Pedidos'...")
            page.locator("span.sub-menu-title", has_text="Pedidos").click()
            page.locator("a[title='Rastreio de pedidos']").click()

            print("INFO: Abrindo a tela de 'Exportar Pedido Avançado'...")
            page.locator("button", has_text="Exportar").hover()
            page.locator("div.ssc-dropdown-item", has_text="Exportar Pedido Avançado").click()

            print("INFO: Selecionando os status de inventário...")
            # Status que representam pacotes que ainda não foram finalizados (em posse da transportadora)
            status_inventario = ["Pickup_Pending", "Pickup_Done", "Hub_Received", "Hub_Assigned", "Hub_Assigning", "Delivering", "OnHold", "Failed Delivery"]
            
            for status in status_inventario:
                # Localiza a linha do status e clica no checkbox correspondente
                linha_status = page.locator("div.s-tree-node__content").filter(
                    has=page.locator(f"span.ssc-tree-node__label:text-is('{status}')")
                )
                if linha_status.count() > 0:
                    linha_status.locator("label.ssc-checkbox-wrapper").click()
                else:
                    print(f"WARN: Status '{status}' não encontrado na tela. Continuando...")
            
            # Garante que "All" em "Conta do pedido" está selecionado
            page.locator("div.ssc-form-item", has_text="Conta do pedido:").locator("span.ssc-tree-node__label", has_text="All").first.click()

            print("INFO: Confirmando a exportação...")
            page.locator("button.ssc-btn-type-primary", has_text="Confirmar").click()

            print("INFO: Aguardando o relatório ser gerado pelo servidor (pode levar alguns minutos)...")
            # Localiza a primeira linha de tarefa "Forward Order" e espera o botão de baixar
            tarefa_row = page.locator("div.task-row", has_text="Forward Order").first
            botao_baixar = tarefa_row.locator("button", has_text="Baixar")
            
            # Timeout estendido para 5 minutos para a Shopee processar
            botao_baixar.wait_for(state="visible", timeout=300000) 
            
            print("INFO: Baixando o arquivo...")
            with page.expect_download() as download_info:
                botao_baixar.click()
                time.sleep(5) # Pequena espera para o download iniciar de fato
            
            download = download_info.value
            caminho_zip = os.path.join(DOWNLOAD_DIR, download.suggested_filename)
            download.save_as(caminho_zip)
            print(f"SUCCESS: Arquivo '{download.suggested_filename}' baixado.")
            
            browser.close()

            if caminho_zip.endswith('.zip'):
                print(f"INFO: Extraindo arquivo ZIP...")
                with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
                    arquivos_extraidos = zip_ref.namelist()
                    zip_ref.extractall(DOWNLOAD_DIR)
                
                # Procura por csv ou xlsx dentro do que foi extraído
                for ext in ["*.csv", "*.xlsx"]:
                    arquivos = glob.glob(os.path.join(DOWNLOAD_DIR, ext))
                    if arquivos:
                        return max(arquivos, key=os.path.getctime)

            return caminho_zip # Retorna o zip se não for um arquivo de planilha

    except Exception as e:
        print(f"FATAL: Ocorreu um erro durante a automação do navegador: {e}")
        return None

# ==========================================
# 2. FUNÇÃO PARA ATUALIZAR AS PLANILHAS
# ==========================================
def atualizar_planilha_latencia(df_completo, hub_filtro, caminho_planilha):
    """
    Filtra o DataFrame por hub e atualiza a Sheet1 da planilha de latência correspondente.
    """
    print(f"INFO: Processando dados para o hub: '{hub_filtro}'...")
    
    # Encontra a coluna da estação atual de forma robusta
    col_station = next((c for c in df_completo.columns if 'current station' in c.lower()), None)
    if not col_station:
        print("FATAL: Coluna 'Current Station' não encontrada no arquivo baixado. Abortando.")
        return

    # Filtra o dataframe para o hub específico (JML ou ITR)
    df_hub = df_completo[df_completo[col_station].str.contains(hub_filtro, case=False, na=False)].copy()

    if df_hub.empty:
        print(f"WARN: Nenhum dado encontrado para o hub '{hub_filtro}' no arquivo de inventário.")
        # Mesmo se estiver vazio, vamos sobrescrever a planilha para limpá-la.
    
    print(f"INFO: {len(df_hub)} registros encontrados para '{hub_filtro}'.")
    
    try:
        print(f"INFO: Atualizando a 'Sheet1' do arquivo '{os.path.basename(caminho_planilha)}'...")
        # Usa o ExcelWriter para substituir apenas a Sheet1, preservando outras abas
        with pd.ExcelWriter(caminho_planilha, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df_hub.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Formata waybill como texto
        formatar_waybill_como_texto(caminho_planilha)
        print(f"SUCCESS: Planilha '{os.path.basename(caminho_planilha)}' foi atualizada com sucesso.")

    except FileNotFoundError:
        print(f"WARN: Arquivo '{os.path.basename(caminho_planilha)}' não encontrado. Criando um novo...")
        df_hub.to_excel(caminho_planilha, sheet_name='Sheet1', index=False)
        formatar_waybill_como_texto(caminho_planilha)
        print(f"SUCCESS: Novo arquivo '{os.path.basename(caminho_planilha)}' criado com os dados.")
        
    except Exception as e:
        print(f"FATAL: Falha ao escrever na planilha '{os.path.basename(caminho_planilha)}': {e}")

# ==========================================
# 3. FUNÇÃO PARA CONSOLIDAR E CALCULAR AGING
# ==========================================
def consolidar_e_calcular_aging():
    """
    Consolida as planilhas JML e ITR, calcula o aging e salva o resultado.
    """
    print("\nINFO: Iniciando a consolidação e o cálculo de aging...")
    CONSOLIDADA_PATH = os.path.join(os.getcwd(), "latencia_consolidada.xlsx")
    
    try:
        # Lê a Sheet1 de cada planilha de latência
        df_jml = pd.read_excel(LATENCIA_JML_PATH, sheet_name='Sheet1', dtype=str)
        df_itr = pd.read_excel(LATENCIA_ITR_PATH, sheet_name='Sheet1', dtype=str)

        # Consolida os dois DataFrames
        df_consolidado = pd.concat([df_jml, df_itr], ignore_index=True)
        print(f"INFO: Dados de JML e ITR consolidados. Total de {len(df_consolidado)} registros.")

        # --- Lógica de Cálculo de Aging ---
        target_col_name = 'Last scan time'
        if target_col_name not in df_consolidado.columns:
            print(f"WARN: A coluna '{target_col_name}' não foi encontrada na planilha consolidada. Pulando o cálculo de aging.")
            df_consolidado.to_excel(CONSOLIDADA_PATH, sheet_name='Sheet1', index=False)
            formatar_waybill_como_texto(CONSOLIDADA_PATH)
            print(f"SUCCESS: Arquivo consolidado '{os.path.basename(CONSOLIDADA_PATH)}' salvo sem a coluna 'aging'.")
            return

        # Define a data de hoje para o cálculo
        hoje = pd.to_datetime("today") # Usar a data atual do sistema

        # Converte a coluna para datetime e calcula o aging
        df_consolidado[target_col_name] = pd.to_datetime(df_consolidado[target_col_name], errors='coerce')
        df_consolidado['aging'] = (hoje - df_consolidado[target_col_name]).dt.days
        
        # Remove a coluna original e renomeia
        df_consolidado.drop(columns=[target_col_name], inplace=True)
        
        print("INFO: Coluna 'aging' calculada com sucesso.")

        # Salva o resultado final na Sheet1 do arquivo consolidado, sobrescrevendo-o
        df_consolidado.to_excel(CONSOLIDADA_PATH, sheet_name='Sheet1', index=False)
        
        # Formata waybill como texto
        formatar_waybill_como_texto(CONSOLIDADA_PATH)
        print(f"SUCCESS: Arquivo '{os.path.basename(CONSOLIDADA_PATH)}' foi atualizado com a coluna 'aging' (waybill em formato texto).")

    except FileNotFoundError as e:
        print(f"FATAL: Um dos arquivos de latência não foi encontrado para consolidação: {e}")
    except Exception as e:
        print(f"FATAL: Ocorreu um erro durante a consolidação e cálculo de aging: {e}")



# ==========================================
# SCRIPT PRINCIPAL
# ==========================================
if __name__ == "__main__":
    # Passo 1: Baixar o relatório de inventário
    caminho_arquivo_inventario = baixar_inventario_shopee()
    
    if caminho_arquivo_inventario:
        # Passo 2: Ler os dados baixados
        df_inventario = ler_planilha_shopee_segura(caminho_arquivo_inventario)
        
        if not df_inventario.empty:
            # Passo 3: Atualizar as planilhas de JML e ITR
            atualizar_planilha_latencia(df_inventario, hub_filtro='JML|Monlevade', caminho_planilha=LATENCIA_JML_PATH)
            atualizar_planilha_latencia(df_inventario, hub_filtro='ITR|Itabira', caminho_planilha=LATENCIA_ITR_PATH)
            
            print("\nINFO: Processo concluído.")
        else:
            print("ERROR: O DataFrame do inventário está vazio. Não foi possível continuar.")
    else:
        print("ERROR: O download do inventário falhou. Não foi possível continuar.")
