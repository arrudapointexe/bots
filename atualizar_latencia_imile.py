import os
import pandas as pd
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv
from openpyxl import load_workbook

# Carrega as variáveis de ambiente do arquivo .env ANTES de qualquer outra coisa
load_dotenv()

# Meus módulos locais, baseados nos seus scripts existentes
from imile_utils import login_imile
from config import BASES_SLA

# ==========================================
# CONFIGURAÇÕES
# ==========================================
# Mapeia a sigla da base para o MESMO arquivo de latência final consolidado
MAPA_ARQUIVOS_LATENCIA = {
    "JML": "latencia_consolidada.xlsx",
    "ITR": "latencia_consolidada.xlsx"
}

# ==========================================
# FUNÇÃO AUXILIAR PARA FORMATAÇÃO
# ==========================================
def formatar_waybill_como_texto(caminho_arquivo):
    """
    Formata todas as colunas de waybill/tracking como texto no Excel para evitar notação científica.
    """
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        # Itera pelas colunas do header para identificar waybill/tracking
        for col_idx, col_cell in enumerate(ws[1], 1):
            if col_cell.value and any(x in str(col_cell.value).lower() for x in ['waybill', 'tracking', 'order', 'pack']):
                # Formata todas as células dessa coluna como texto
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = '@'  # @ = texto no Excel
                        if cell.value is not None:
                            cell.value = str(cell.value)  # Converte para string
        
        wb.save(caminho_arquivo)
        print(f"INFO: Arquivo '{os.path.basename(caminho_arquivo)}' formatado com waybill como texto.")
    except Exception as e:
        print(f"WARN: Erro ao formatar arquivo: {e}")

# ==========================================
# FUNÇÃO DE DOWNLOAD
# ==========================================
def baixar_inventario_imile(page, sigla):
    """
    Navega no portal iMile e baixa o relatório de inventário para a base (sigla) logada.
    Retorna o caminho do arquivo temporário baixado.
    """
    print(f"[{sigla}] INFO: Navegando para a tela de 'Inventory Monitor'...")
    
    # Usamos try/except para cada clique para dar mais robustez e logs claros
    try:
        page.get_by_text("Monitor").first.click(force=True, timeout=20000)
    except Exception as e:
        print(f"[{sigla}] ERROR: Não foi possível clicar no menu 'Monitor'. Tentando alternativa... Detalhe: {e}")
        # Se o menu principal falha, às vezes um F5 e tentar de novo ajuda.
        page.reload(wait_until="domcontentloaded")
        page.get_by_text("Monitor").first.click(force=True, timeout=20000)
        
    page.wait_for_timeout(2000) # Pausa curta para o submenu abrir
    
    try:
        page.get_by_text("Operation Monitor").first.click(force=True)
    except Exception:
        # Tenta pelo link direto se o menu não funcionar
        print(f"[{sigla}] WARN: Clique em 'Operation Monitor' falhou, tentando acessar pelo URL.")
        current_url = page.url
        base_url = current_url.split('/#/') [0]
        page.goto(f"{base_url}/#/operation-monitor/inventory-monitor", wait_until="domcontentloaded")


    page.wait_for_timeout(2000)
    
    try:
        page.get_by_text("Inventory Monitor").first.click(force=True)
    except Exception:
        print(f"[{sigla}] WARN: Clique em 'Inventory Monitor' falhou, assumindo que já estamos na página correta.")
    
    # Espera a página de inventário carregar um elemento chave
    print(f"[{sigla}] INFO: Aguardando a página de inventário carregar...")
    page.locator('.ImileActionButton-root:has-text("Export")').first.wait_for(state="visible", timeout=45000)

    print(f"[{sigla}] INFO: Solicitando a exportação do inventário completo...")
    page.locator('.ImileActionButton-root:has-text("Export")').first.click(force=True)
    page.wait_for_timeout(1000)
    page.locator('li.ImileMenuItem-root:has-text("Export All")').first.click(force=True)
    page.wait_for_timeout(1000)
    
    # Clica no botão de confirmação dentro do modal de exportação, usando o seletor de classe mais robusto
    page.locator('.export-button').first.click(force=True)

    print(f"[{sigla}] INFO: Aguardando a iMile gerar o arquivo (isso pode demorar)...")
    page.wait_for_timeout(15000) # Espera estendida para a geração

    print(f"[{sigla}] INFO: Abrindo a central de downloads para baixar o arquivo...")
    # O ícone de download na barra de navegação superior
    page.locator('span.Imile-ButtonIcon-root svg path[d*="M8 0C3.57"]').locator('..').first.click(force=True)

    # Espera o último botão de download aparecer na lista
    btn_baixar = page.locator('button:has-text("download")').last
    btn_baixar.wait_for(state="visible", timeout=30000)

    # Inicia o download
    print(f"[{sigla}] INFO: Baixando o arquivo de inventário...")
    with page.expect_download(timeout=120000) as download_info:
        btn_baixar.click(force=True)

    download = download_info.value
    # Salva com um nome temporário para não confundir com arquivos finais
    caminho_temporario = f"temp_inventario_{sigla}.xlsx"
    download.save_as(caminho_temporario)

    print(f"[{sigla}] SUCCESS: Inventário baixado e salvo como '{caminho_temporario}'")
    return caminho_temporario

# ==========================================
# FUNÇÃO PARA PREPARAR O DATAFRAME DE UMA BASE
# ==========================================
def preparar_dataframe_base(caminho_origem, sigla_base):
    """
    Lê um arquivo de inventário, adiciona a coluna 'Base' e o retorna.
    Não faz nenhuma operação de escrita em arquivo.
    """
    if not os.path.exists(caminho_origem):
        print(f"ERROR: O arquivo de origem '{caminho_origem}' não foi encontrado.")
        return None

    try:
        print(f"INFO: Lendo dados do novo inventário de '{caminho_origem}'...")
        df_novo = pd.read_excel(caminho_origem)
        
        df_novo['Base'] = sigla_base
        
        colunas = ['Base'] + [col for col in df_novo.columns if col != 'Base']
        df_novo = df_novo[colunas]
        
        print(f"INFO: DataFrame para a base {sigla_base} preparado na memória.")
        return df_novo
        
    except Exception as e:
        print(f"ERROR: Falha ao processar o arquivo Excel de origem '{caminho_origem}': {e}")
        return None

# ==========================================
# FUNÇÃO FINAL PARA CONSOLIDAR, CALCULAR AGING E SALVAR
# ==========================================
def consolidar_calcular_e_salvar(lista_dfs, caminho_destino):
    """
    Recebe uma lista de DataFrames, consolida, calcula o aging e salva o resultado final.
    """
    if not lista_dfs:
        print("WARN: Nenhum dado foi coletado para processar. O arquivo final não será gerado.")
        return

    print("\nINFO: Etapa final: Consolidando todos os dados...")
    df_final = pd.concat(lista_dfs, ignore_index=True)
    print(f"INFO: Consolidação concluída. Total de {len(df_final)} registros.")

    nome_aba_dados = "Sheet1"
    target_col_name = 'Last scan time'

    # --- Cálculo de Aging ---
    if target_col_name not in df_final.columns:
        print(f"WARN: A coluna '{target_col_name}' não foi encontrada. O 'aging' não será calculado.")
    else:
        print("INFO: Calculando 'aging'...")
        hoje = pd.to_datetime("today").normalize()
        
        # Converte a coluna alvo para datetime e normaliza para meia-noite (corrige o bug do -1)
        df_final[target_col_name] = pd.to_datetime(df_final[target_col_name], errors='coerce').dt.normalize()
        
        # Calcula a diferença em dias
        aging_values = (hoje - df_final[target_col_name]).dt.days
        
        # Renomeia a coluna e preenche com os valores calculados
        df_final.rename(columns={target_col_name: 'aging'}, inplace=True)
        df_final['aging'] = aging_values
        print("INFO: Coluna 'aging' calculada com sucesso.")

    # --- Salvamento Final ---
    try:
        print(f"INFO: Salvando resultado final em '{caminho_destino}', aba '{nome_aba_dados}'...")
        
        # Salva como texto para evitar notação científica
        with pd.ExcelWriter(caminho_destino, engine='openpyxl', mode='w') as writer:
            df_final.to_excel(writer, sheet_name=nome_aba_dados, index=False)
        
        # Formata waybill como texto
        formatar_waybill_como_texto(caminho_destino)
        print(f"SUCCESS: Planilha '{caminho_destino}' gerada com sucesso (waybill em formato texto).")

    except PermissionError:
        print(f"FATAL: PERMISSÃO NEGADA para escrever no arquivo '{caminho_destino}'. Certifique-se de fechar o Excel.")
    except Exception as e:
        print(f"FATAL: Falha inesperada ao salvar a planilha final: {e}")


# ==========================================
# SCRIPT PRINCIPAL
# ==========================================
def main():
    print("======================================================")
    print("🚀 INICIANDO ATUALIZADOR DE LATÊNCIA IMILE 🚀")
    print("======================================================")
    
    lista_dataframes = [] # Lista para guardar os dataframes de cada base

    with sync_playwright() as p:
        for sigla, usuario, senha in BASES_SLA:
            caminho_temporario = None
            print(f"\n--- Processando base: {sigla} ---")
            
            if not usuario or not senha:
                print(f"WARN: Credenciais para '{sigla}' não encontradas no .env. Pulando...")
                continue

            browser = p.chromium.launch(headless=True, args=["--start-maximized"])
            context = browser.new_context(no_viewport=True, accept_downloads=True)
            page = context.new_page()

            try:
                login_imile(page, usuario, senha)
                caminho_temporario = baixar_inventario_imile(page, sigla)

                if caminho_temporario:
                    df_base = preparar_dataframe_base(caminho_temporario, sigla)
                    if df_base is not None:
                        lista_dataframes.append(df_base)
                else:
                    print(f"ERROR: Download para a base '{sigla}' falhou.")

            except Exception as e:
                print(f"FATAL: Ocorreu um erro crítico no processamento da base {sigla}: {e}")
                page.screenshot(path=f"error_{sigla}.png")
                print(f"INFO: Um print do erro foi salvo como 'error_{sigla}.png'")

            finally:
                print(f"[{sigla}] INFO: Finalizando e limpando...")
                context.close()
                browser.close()
                if caminho_temporario and os.path.exists(caminho_temporario):
                    os.remove(caminho_temporario)
                    print(f"[{sigla}] INFO: Arquivo temporário '{caminho_temporario}' removido.")
        
    # --- Etapa Final: Consolidar, Calcular Aging e Salvar ---
    if BASES_SLA:
        primeira_sigla = BASES_SLA[0][0]
        arquivo_consolidado_final = MAPA_ARQUIVOS_LATENCIA.get(primeira_sigla)
        if arquivo_consolidado_final:
             consolidar_calcular_e_salvar(lista_dataframes, arquivo_consolidado_final)

    print("\n======================================================")
    print("✅ Processo concluído para todas as bases.")
    print("======================================================")


if __name__ == "__main__":
    main()